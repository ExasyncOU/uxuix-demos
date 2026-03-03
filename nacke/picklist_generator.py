"""
picklist_generator.py – AB-Picklisten als Excel generieren.

Drei Template-Varianten (basierend auf echten Nacke AB-Dateien):

1. PCS (Picking=1): Mo/Do-Split, Bundle Anzahl, korrigierte Menge pro Land
2. PPP (Picking=3): Größe+Packs-Paare, Abruf-Spalte, KEIN Mo/Do-Split
3. SUP (Picking=2): PCS-Layout, aber wenig Größen

Gruppiert Post-Over-Export nach AROS-Code (Supplier-Class-Code-Serial).
Listennummer: fortlaufend 1001-9999, gespeichert in state/list_counter.json.
Dateiname: KI_UXUIX-[4stellig ListNr]-[OrderCode]-[Datum].xlsx
Ausgabe: K:\\Raussuchlisten\\2026 (konfigurierbar)
"""

import json
import logging
import math
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from master_data import MasterData

logger = logging.getLogger(__name__)

# Excel-Stile
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13)
BUNDLE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
SUM_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")

# Länderkürzel-Reihenfolge (wie in den Referenz-ABs)
COUNTRY_ORDER = ["NL", "D", "B", "F", "CH", "E", "A", "SK", "OL"]


class PicklistGenerator:
    """Generiert AB-Picklisten aus Post-Overconfirmation-Export."""

    def __init__(self, master: MasterData, config_path: str = "config.json"):
        self.master = master
        with open(config_path, "r", encoding="utf-8") as f:
            self.cfg = json.load(f)

        self.output_dir = Path(self.cfg["paths"]["output_dir"])
        self.state_dir = Path(self.cfg["paths"]["state_dir"])
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.state_dir.mkdir(parents=True, exist_ok=True)

        pl_cfg = self.cfg.get("picklist", {})
        self.list_min = pl_cfg.get("list_number_min", 1001)
        self.list_max = pl_cfg.get("list_number_max", 9999)

    # ------------------------------------------------------------------
    # Hauptmethode
    # ------------------------------------------------------------------

    def generate(self, export_rows: list[dict], run_date: date | None = None) -> list[dict]:
        """Generiert alle Picklisten aus dem Export.

        Returns: Liste von Dicts {list_nr, filename, filepath, aros_key, abrufart, picking, total_qty}
        """
        if run_date is None:
            run_date = date.today()

        date_str = run_date.strftime("%d.%m.%y")
        normalized = self._normalize_rows(export_rows)

        # Gruppiere: AROS-Key → Zeilen (eine Liste pro AROS)
        groups = self._group_by_aros(normalized)

        generated = []
        for aros_key, rows in groups.items():
            # Abrufart aus den Rows bestimmen (sollte einheitlich pro AROS sein)
            abrufart = rows[0].get("abrufart", "PCS").upper()
            picking_method = rows[0].get("picking_method", "1")

            list_nr = self._next_list_number()
            order_code = aros_key.replace("-", ".")

            suffix = f" ({abrufart})" if abrufart != "PCS" else ""
            filename = f"KI_UXUIX-{list_nr:04d}-{order_code}-{date_str}{suffix}.xlsx".replace("/", "-")
            filepath = self.output_dir / filename

            # Template-Variante wählen
            if abrufart == "PPP":
                total_qty = self._create_ppp_picklist(filepath, aros_key, rows, list_nr, date_str)
            else:
                total_qty = self._create_pcs_picklist(filepath, aros_key, abrufart, rows, list_nr, date_str)

            generated.append({
                "list_nr": list_nr,
                "filename": filename,
                "filepath": str(filepath),
                "aros_key": aros_key,
                "abrufart": abrufart,
                "picking": "Mixed",
                "total_qty": total_qty,
            })
            logger.info("Liste %04d: %s %s, %d Stk -> %s", list_nr, aros_key, abrufart, total_qty, filename)

        logger.info("Insgesamt %d Picklisten generiert.", len(generated))
        return generated

    # ------------------------------------------------------------------
    # PCS-Template (mit Mo/Do-Split, Bundle Anzahl, korrigierte Menge)
    # ------------------------------------------------------------------

    def _create_pcs_picklist(self, filepath, aros_key, abrufart, rows, list_nr, date_str) -> int:
        """Erstellt PCS/SUP-Picklist im AB-Format.

        Layout (wie 9004/9020 Referenz):
          Row 1-6: Header-Block
          Row 7: Spaltenheader: Order-Nr. | Land | Size1 | Size2 | ... | SUMME | Store/Depot | ITEM-Nr. | PREIS
          Row 8+: Pro Land-Block:
            - Datenzeile
            - "Bundle Anzahl"
            - "korrigierte Menge (Bundle)"
            - "Gesamt keine Info:"
          Footer: Gesamt Stores ohne D, Gesamt Stores, Gesamt ALLES
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Uebersicht"

        # Sammle alle vorhandenen Länder und Größen
        countries = self._sort_countries(set(r["country"] for r in rows))
        sizes = self._collect_sizes(rows)

        # Pivot: Land → Größe → {qty, item_number, price, order_type}
        pivot = self._build_pivot(rows)

        # Bundle Size aus Masterdatei
        order_info = self.master.get_order_info(aros_key)
        bundle_store = order_info["bundle_size_store"] if order_info else 0
        bundle_depot = order_info["bundle_size_depot"] if order_info else 0
        ol_bundle = self.master.get_ol_bundle(aros_key) or 0

        # --- Header-Block (Rows 1-6) ---
        ws.cell(row=1, column=1, value=f"Order-Nr.: {aros_key}").font = TITLE_FONT
        ws.cell(row=2, column=1, value=f"Liste: KI_UXUIX-{list_nr:04d}").font = Font(bold=True, size=11)
        ws.cell(row=3, column=1, value=f"Datum: {date_str}").font = Font(size=10)
        ws.cell(row=4, column=1, value=f"Abrufart: {abrufart}").font = Font(size=10)
        if order_info:
            ws.cell(row=5, column=1, value=f"Artikel: {order_info.get('artikel', '')}").font = Font(size=10)

        # --- Row 7: Spaltenheader ---
        header_row = 7
        headers = ["Order-Nr.", "Land"] + [str(s) for s in sizes] + ["SUMME", "Store/Depot", "ITEM-Nr.", "PREIS"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = CENTER

        # --- Land-Blöcke ---
        current_row = header_row + 1
        total_all = 0
        total_stores = 0
        total_stores_no_d = 0

        for country in countries:
            if country not in pivot:
                continue

            country_data = pivot[country]
            country_total = 0

            # Datenzeile
            ws.cell(row=current_row, column=1, value=aros_key).border = THIN_BORDER
            ws.cell(row=current_row, column=2, value=country).border = THIN_BORDER
            ws.cell(row=current_row, column=2).font = Font(bold=True)

            for si, size in enumerate(sizes, 3):
                qty = country_data.get(str(size), {}).get("qty", 0)
                cell = ws.cell(row=current_row, column=si)
                if qty > 0:
                    cell.value = qty
                cell.border = THIN_BORDER
                cell.alignment = CENTER
                country_total += qty

            # Summe, Store/Depot, ITEM, PREIS
            sum_col = len(sizes) + 3
            ws.cell(row=current_row, column=sum_col, value=country_total).border = THIN_BORDER
            ws.cell(row=current_row, column=sum_col).font = Font(bold=True)
            ws.cell(row=current_row, column=sum_col).alignment = CENTER

            # Store/Depot, Item, Preis aus erster Zeile des Landes
            first_entry = next(iter(country_data.values()), {})
            order_type = first_entry.get("order_type", "S")
            ws.cell(row=current_row, column=sum_col + 1, value=order_type).border = THIN_BORDER
            ws.cell(row=current_row, column=sum_col + 2, value=first_entry.get("item_number", "")).border = THIN_BORDER
            ws.cell(row=current_row, column=sum_col + 3, value=first_entry.get("price", "")).border = THIN_BORDER

            current_row += 1

            # Bundle Anzahl Zeile
            bs = bundle_store if order_type == "S" else bundle_depot
            if country.upper() == "OL" and ol_bundle > 0:
                bs = ol_bundle

            ws.cell(row=current_row, column=2, value="Bundle Anzahl").font = Font(italic=True)
            ws.cell(row=current_row, column=2).fill = BUNDLE_FILL
            ws.cell(row=current_row, column=2).border = THIN_BORDER
            bundle_total = 0
            for si, size in enumerate(sizes, 3):
                qty = country_data.get(str(size), {}).get("qty", 0)
                cell = ws.cell(row=current_row, column=si)
                cell.fill = BUNDLE_FILL
                cell.border = THIN_BORDER
                cell.alignment = CENTER
                if bs > 0 and qty > 0:
                    bundles = math.ceil(qty / bs)
                    cell.value = bundles
                    bundle_total += bundles
            ws.cell(row=current_row, column=sum_col, value=bundle_total if bs > 0 else "").fill = BUNDLE_FILL
            ws.cell(row=current_row, column=sum_col).border = THIN_BORDER
            current_row += 1

            # korrigierte Menge (Bundle) Zeile
            label = "korrigierte Menge (Krt.)" if country.upper() == "OL" else "korrigierte Menge (Bundle)"
            ws.cell(row=current_row, column=2, value=label).font = Font(italic=True)
            ws.cell(row=current_row, column=2).border = THIN_BORDER
            corrected_total = 0
            for si, size in enumerate(sizes, 3):
                qty = country_data.get(str(size), {}).get("qty", 0)
                cell = ws.cell(row=current_row, column=si)
                cell.border = THIN_BORDER
                cell.alignment = CENTER
                if bs > 0 and qty > 0:
                    corrected = math.ceil(qty / bs) * int(bs)
                    cell.value = corrected
                    corrected_total += corrected
                elif qty > 0:
                    cell.value = qty
                    corrected_total += qty
            ws.cell(row=current_row, column=sum_col, value=corrected_total).border = THIN_BORDER
            ws.cell(row=current_row, column=sum_col).font = Font(bold=True)
            current_row += 1

            # Gesamt-Zeile pro Land
            ws.cell(row=current_row, column=2, value="Gesamt keine Info:").font = Font(bold=True, size=10)
            ws.cell(row=current_row, column=2).fill = SUM_FILL
            ws.cell(row=current_row, column=2).border = THIN_BORDER
            ws.cell(row=current_row, column=sum_col, value=country_total).fill = SUM_FILL
            ws.cell(row=current_row, column=sum_col).border = THIN_BORDER
            ws.cell(row=current_row, column=sum_col).font = Font(bold=True)
            current_row += 1

            total_all += country_total
            if order_type == "S":
                total_stores += country_total
                if country.upper() != "D":
                    total_stores_no_d += country_total

            current_row += 1  # Leerzeile zwischen Ländern

        # --- Footer ---
        current_row += 1
        ws.cell(row=current_row, column=2, value="Gesamt Stores ohne D").font = Font(bold=True)
        ws.cell(row=current_row, column=2).fill = SUM_FILL
        ws.cell(row=current_row, column=len(sizes) + 3, value=total_stores_no_d).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=len(sizes) + 3).fill = SUM_FILL
        current_row += 1

        ws.cell(row=current_row, column=2, value="Gesamt Stores").font = Font(bold=True)
        ws.cell(row=current_row, column=2).fill = SUM_FILL
        ws.cell(row=current_row, column=len(sizes) + 3, value=total_stores).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=len(sizes) + 3).fill = SUM_FILL
        current_row += 1

        ws.cell(row=current_row, column=2, value="Gesamt ALLES").font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=2).fill = SUM_FILL
        ws.cell(row=current_row, column=len(sizes) + 3, value=total_all).font = Font(bold=True, size=14)
        ws.cell(row=current_row, column=len(sizes) + 3).fill = SUM_FILL

        # Spaltenbreiten
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 28
        for ci in range(3, len(sizes) + 7):
            ws.column_dimensions[get_column_letter(ci)].width = 10
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1

        # Daten-Sheet: Rohdaten aus Export (Spalten M/N/O/Q/R/S/T)
        self._create_data_sheet(wb, aros_key, rows)

        wb.save(filepath)
        return total_all

    # ------------------------------------------------------------------
    # PPP-Template (Größe+Packs-Paare, kein Mo/Do-Split)
    # ------------------------------------------------------------------

    def _create_ppp_picklist(self, filepath, aros_key, rows, list_nr, date_str) -> int:
        """Erstellt PPP-Picklist im AB-Format.

        Layout (wie 9135 Referenz):
          Row 7: Order-Nr. | Land | (blank) | Abruf | Size1 | Packs Gr.Size1 | Size2 | ... | SUMME | S/D | ITEM | PREIS
          Pro Land: eine Datenzeile + Gesamt
          Kein Mo/Do-Split, keine Bundle-Zeilen
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Uebersicht"

        countries = self._sort_countries(set(r["country"] for r in rows))
        sizes = self._collect_sizes(rows)
        pivot = self._build_pivot(rows)

        order_info = self.master.get_order_info(aros_key)
        bundle_store = order_info["bundle_size_store"] if order_info else 0
        bundle_depot = order_info["bundle_size_depot"] if order_info else 0
        ol_bundle = self.master.get_ol_bundle(aros_key) or 0

        # Header
        ws.cell(row=1, column=1, value=f"Order-Nr.: {aros_key}").font = TITLE_FONT
        ws.cell(row=2, column=1, value=f"Liste: KI_UXUIX-{list_nr:04d}  (PPP)").font = Font(bold=True, size=11)
        ws.cell(row=3, column=1, value=f"Datum: {date_str}").font = Font(size=10)

        # Row 7: Headers – Size-Spalten kommen in Paaren: Qty + "Packs Gr. X"
        header_row = 7
        headers = ["Order-Nr.", "Land", "", "Abruf"]
        for s in sizes:
            headers.append(str(s))
            headers.append(f"Packs Gr. {s}")
        headers += ["SUMME", "Store/Depot", "ITEM-Nr.", "PREIS"]

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = CENTER

        current_row = header_row + 1
        total_all = 0

        for country in countries:
            if country not in pivot:
                continue

            country_data = pivot[country]
            country_total = 0

            ws.cell(row=current_row, column=1, value=aros_key).border = THIN_BORDER
            ws.cell(row=current_row, column=2, value=country).border = THIN_BORDER
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            # Col 4 = Abruf (Allocation number – from order data)
            ws.cell(row=current_row, column=4, value="PPP").border = THIN_BORDER

            # Bundle Size: OL-Bundle für OL, sonst Store/Depot je nach order_type
            first_entry = next(iter(country_data.values()), {})
            order_type = first_entry.get("order_type", "S")
            if country.upper() == "OL" and ol_bundle > 0:
                bs = ol_bundle
            elif order_type == "D":
                bs = bundle_depot
            else:
                bs = bundle_store

            col_offset = 5
            for size in sizes:
                qty = country_data.get(str(size), {}).get("qty", 0)
                # Qty
                cell = ws.cell(row=current_row, column=col_offset)
                if qty > 0:
                    cell.value = qty
                cell.border = THIN_BORDER
                cell.alignment = CENTER
                # Packs
                packs_cell = ws.cell(row=current_row, column=col_offset + 1)
                if qty > 0 and bs > 0:
                    packs_cell.value = math.ceil(qty / bs)
                packs_cell.border = THIN_BORDER
                packs_cell.alignment = CENTER
                country_total += qty
                col_offset += 2

            # SUMME, S/D, ITEM, PREIS
            ws.cell(row=current_row, column=col_offset, value=country_total).border = THIN_BORDER
            ws.cell(row=current_row, column=col_offset).font = Font(bold=True)

            ws.cell(row=current_row, column=col_offset + 1, value=order_type).border = THIN_BORDER
            ws.cell(row=current_row, column=col_offset + 2, value=first_entry.get("item_number", "")).border = THIN_BORDER
            ws.cell(row=current_row, column=col_offset + 3, value=first_entry.get("price", "")).border = THIN_BORDER

            current_row += 1
            total_all += country_total

        # Gesamt ALLES
        current_row += 1
        ws.cell(row=current_row, column=2, value="Gesamt ALLES").font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=2).fill = SUM_FILL
        sum_col = 5 + len(sizes) * 2
        ws.cell(row=current_row, column=sum_col, value=total_all).font = Font(bold=True, size=14)
        ws.cell(row=current_row, column=sum_col).fill = SUM_FILL

        # Spaltenbreiten
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 12
        for ci in range(3, sum_col + 5):
            ws.column_dimensions[get_column_letter(ci)].width = 12
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1

        # Daten-Sheet: Rohdaten aus Export (Spalten M/N/O/Q/R/S/T)
        self._create_data_sheet(wb, aros_key, rows)

        wb.save(filepath)
        return total_all

    # ------------------------------------------------------------------
    # Hilfsfunktionen
    # ------------------------------------------------------------------

    def _create_data_sheet(self, wb: Workbook, aros_key: str, rows: list[dict]) -> None:
        """Fügt 'Daten' Sheet mit Rohdaten des Exports hinzu.

        Enthält die Originaldaten aus dem IDIS-Export:
        Spalten entsprechen M/N/O/Q/R/S/T des CSV-Exports.
        """
        ws = wb.create_sheet(title="Daten")
        headers = ["Ordernummer", "Land", "Größe", "Stückzahl", "Item Nr", "Preis", "S/D"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = CENTER

        sorted_rows = sorted(rows, key=lambda x: (x.get("country", ""), x.get("size", "")))
        for ri, r in enumerate(sorted_rows, 2):
            ws.cell(row=ri, column=1, value=r.get("order_id", aros_key)).border = THIN_BORDER
            ws.cell(row=ri, column=2, value=r.get("country", "")).border = THIN_BORDER
            ws.cell(row=ri, column=3, value=r.get("size", "")).border = THIN_BORDER
            qty_cell = ws.cell(row=ri, column=4, value=r.get("quantity", 0))
            qty_cell.border = THIN_BORDER
            qty_cell.alignment = CENTER
            ws.cell(row=ri, column=5, value=r.get("item_number", "")).border = THIN_BORDER
            ws.cell(row=ri, column=6, value=r.get("selling_price", "")).border = THIN_BORDER
            sd_cell = ws.cell(row=ri, column=7, value=r.get("order_type", ""))
            sd_cell.border = THIN_BORDER
            sd_cell.alignment = CENTER

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 6

    @staticmethod
    def _sort_countries(countries: set) -> list:
        """Sortiert Länder in der Nacke-Standardreihenfolge."""
        order_map = {c: i for i, c in enumerate(COUNTRY_ORDER)}
        return sorted(countries, key=lambda c: order_map.get(c.upper(), 99))

    # Bekleidungsgrößen-Reihenfolge
    _SIZE_ORDER = {"XXS": 0, "XS": 1, "S": 2, "M": 3, "L": 4, "XL": 5, "XXL": 6, "3XL": 7}

    @classmethod
    def _collect_sizes(cls, rows: list[dict]) -> list:
        """Sammelt alle Größen und sortiert: numerisch oder nach Bekleidungsgrößen."""
        sizes = []
        seen = set()
        for r in rows:
            s = str(r.get("size", "")).strip()
            if s and s not in seen:
                sizes.append(s)
                seen.add(s)

        def size_key(s):
            upper = s.upper()
            if upper in cls._SIZE_ORDER:
                return (0, cls._SIZE_ORDER[upper])
            try:
                return (1, int(s))
            except ValueError:
                return (2, s)

        sizes.sort(key=size_key)
        return sizes

    @staticmethod
    def _build_pivot(rows: list[dict]) -> dict:
        """Baut Pivot: Land → Größe → {qty, item_number, price, order_type}."""
        pivot = {}
        for r in rows:
            country = r.get("country", "").upper()
            size = str(r.get("size", ""))
            qty = r.get("quantity", 0)

            if country not in pivot:
                pivot[country] = {}
            if size not in pivot[country]:
                pivot[country][size] = {
                    "qty": 0,
                    "item_number": r.get("item_number", ""),
                    "price": r.get("selling_price", r.get("price", "")),
                    "order_type": r.get("order_type", r.get("picking", "S")),
                }
            pivot[country][size]["qty"] += qty

        return pivot

    @staticmethod
    def _group_by_aros(rows: list[dict]) -> dict[str, list[dict]]:
        """Gruppiert Zeilen nach AROS-Key."""
        groups = {}
        for row in rows:
            aros = row.get("aros_key", "")
            if aros:
                groups.setdefault(aros, []).append(row)
        return groups

    def _normalize_rows(self, export_rows: list[dict]) -> list[dict]:
        """Normalisiert die Feld-Namen der Export-Zeilen."""
        normalized = []
        for row in export_rows:
            # Spalte T (committed_qty_for_size) bevorzugen – das ist die
            # tatsächlich bestätigte Menge nach Overconfirmation (Post-Over).
            # Fallback auf Spalte P (quantity = qty_for_size) bei Pre-Over-Export.
            qty = row.get("committed_qty_for_size") or row.get("quantity", row.get("menge", 0))
            n = {
                "aros_key": str(row.get("aros_key", row.get("AROS", ""))).strip(),
                "country": str(row.get("country", row.get("land", ""))).strip().upper(),
                "size": str(row.get("size", row.get("groesse", ""))).strip(),
                "quantity": self._to_int(qty),
                "order_id": str(row.get("order_id", row.get("Order", ""))).strip(),
                "abrufart": str(row.get("abrufart", "PCS")).strip().upper(),
                "picking_method": str(row.get("picking_method", "1")).strip(),
                "picking": str(row.get("picking", "Store")).strip(),
                "order_type": str(row.get("order_type", "S")).strip().upper(),
                "item_number": str(row.get("item_number", "")).strip(),
                "selling_price": str(row.get("selling_price", row.get("price", ""))).strip(),
            }

            # Abrufart aus Masterdatei ergänzen wenn fehlend
            if n["abrufart"] in ("", "nan", "NAN"):
                order_info = self.master.get_order_info(n["aros_key"])
                if order_info:
                    n["abrufart"] = order_info["abrufart"].upper()
                else:
                    n["abrufart"] = "PCS"

            if n["aros_key"] and n["quantity"] > 0:
                normalized.append(n)

        return normalized

    # ------------------------------------------------------------------
    # Listennummer-Verwaltung
    # ------------------------------------------------------------------

    def _next_list_number(self) -> int:
        """Liefert die nächste Listennummer (1001-9999, fortlaufend, keine Lücken).

        Startet bei 1001 wenn kein Counter vorhanden.
        Führt tagübergreifend fort – letzten Stand aus Vortag abrufen und weiterführen.
        Verwendet File-Locking um Race Conditions bei parallelen Runs zu vermeiden.
        """
        counter_file = self.state_dir / "list_counter.json"
        lock_file = self.state_dir / "list_counter.lock"

        import msvcrt

        # File-Lock: verhindert doppelte Listennummern bei parallelen Runs
        with open(lock_file, "w") as lf:
            try:
                msvcrt.locking(lf.fileno(), msvcrt.LK_NBLCK, 1)
            except (OSError, IOError):
                # Lock nicht verfügbar – warten und erneut versuchen
                import time
                time.sleep(0.5)
                try:
                    msvcrt.locking(lf.fileno(), msvcrt.LK_NBLCK, 1)
                except (OSError, IOError):
                    logger.warning("List counter lock nicht verfügbar – fahre ohne Lock fort")

            try:
                if counter_file.exists():
                    with open(counter_file, "r", encoding="utf-8") as f:
                        state = json.load(f)
                    current = state.get("current", self.list_min - 1)
                else:
                    state = {}
                    current = self.list_min - 1  # Erster Aufruf → list_min

                next_nr = current + 1
                if next_nr > self.list_max:
                    next_nr = self.list_min

                state["current"] = next_nr
                state["last_updated"] = str(date.today())
                with open(counter_file, "w", encoding="utf-8") as f:
                    json.dump(state, f, indent=2)
            finally:
                try:
                    msvcrt.locking(lf.fileno(), msvcrt.LK_UNLCK, 1)
                except (OSError, IOError):
                    pass

        return next_nr

    def get_current_list_number(self) -> int:
        """Gibt die aktuelle Listennummer zurück."""
        counter_file = self.state_dir / "list_counter.json"
        if counter_file.exists():
            with open(counter_file, "r", encoding="utf-8") as f:
                return json.load(f).get("current", 0)
        return 0

    @staticmethod
    def _to_int(val) -> int:
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return 0


# ------------------------------------------------------------------
# Standalone-Test
# ------------------------------------------------------------------

if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)

    md = MasterData()
    gen = PicklistGenerator(md)

    # Test mit Dummy-PCS-Daten
    test_pcs = [
        {"aros_key": "608-16-901", "country": "AT", "size": "S", "quantity": 12, "abrufart": "PCS", "picking_method": "1", "order_type": "S"},
        {"aros_key": "608-16-901", "country": "AT", "size": "M", "quantity": 24, "abrufart": "PCS", "picking_method": "1", "order_type": "S"},
        {"aros_key": "608-16-901", "country": "D", "size": "S", "quantity": 30, "abrufart": "PCS", "picking_method": "1", "order_type": "S"},
        {"aros_key": "608-16-901", "country": "CH", "size": "L", "quantity": 18, "abrufart": "PCS", "picking_method": "1", "order_type": "S"},
    ]

    result = gen.generate(test_pcs)
    for r in result:
        print(f"  Liste {r['list_nr']:04d}: {r['filename']} ({r['total_qty']} Stk)")
